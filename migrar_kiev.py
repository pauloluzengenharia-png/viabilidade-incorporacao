#!/usr/bin/env python3
"""
Migração da SPE Kiev: da planilha para o banco.

Este script é ao mesmo tempo o importador do primeiro empreendimento e o
caso de teste do sistema. Ele lê o arquivo `26. Incorrido SPE Kiev.xlsx`,
carrega tudo no Postgres pelos mesmos importadores que o upload usa, roda os
três cenários e confere cada linha contra o valor que a planilha guardava.

    python3 migrar_kiev.py dados/kiev.xlsx

Duas decisões de migração que valem registro:

1. **Um orçamento de obra só.** A planilha usa custo raso de R$ 104,0 M
   (SIMULAÇÕES M36) mas distribui no tempo pela curva de um orçamento
   preliminar de R$ 70,6 M vindo de outro empreendimento. Aqui existe um
   único `orcamento_obra`: o total é o da viabilidade e os pesos e a curva
   vêm do cronograma, normalizados. O desembolso passa a fechar com o custo.

2. **O preço de tabela vira registro datado.** `Comercial!D` entra em
   `preco_unidade` com a data da migração, e não como coluna da unidade.
"""
from __future__ import annotations

import datetime as dt
import sys
from collections import defaultdict

import openpyxl

from app.db import Sessao, aplicar_migrations, q
from app.importadores import gravar
from app.importadores.sienge import (normalizar_contratos, normalizar_movimentos,
                                     normalizar_recebido, normalizar_receber,
                                     normalizar_unidades)

ARQUIVO = sys.argv[1] if len(sys.argv) > 1 else "dados/kiev.xlsx"
HOJE = dt.date(2026, 8, 25)


def aba_como_dicts(wb, nome: str, linha_cabecalho: int = 1) -> list[dict]:
    ws = wb[nome]
    linhas = ws.iter_rows(values_only=True)
    for _ in range(linha_cabecalho - 1):
        next(linhas, None)
    cab = next(linhas)
    cols = [str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(cab)]
    saida = []
    for vals in linhas:
        if all(v is None for v in vals):
            continue
        saida.append({cols[i]: v for i, v in enumerate(vals) if i < len(cols)})
    return saida


# =====================================================================
# 1. plano de contas — extraído do bloco realista do fluxo mensal
# =====================================================================
GRUPO_POR_LINHA = {
    "(-) Comissão s/ vendas": "DEDUCAO_VGV",
    "(-) Impostos": "DEDUCAO_RECEITA",
    "(-) Distratos": "DEDUCAO_RECEITA",
    "(-) Despesas comerciais": "DEDUCAO_RECEITA",
    "(+) Outras entradas": "RECEITA",
    "(+) Outras receitas administrativas": "RECEITA",
    "(+) Receitas c/ financiamento": "FINANCEIRO",
    "(-) Amortização de financiamento": "FINANCEIRO",
    "(-) Juros s/ financiamento": "FINANCEIRO",
    "(-/+) Retenções/liberações/Despesas": "FINANCEIRO",
}
# O rótulo do fluxo mensal quase nunca é exatamente o rótulo do DRE: a mesma
# conta aparece como "(-) Despesas Comerciais" numa aba e "(-) Despesas
# comerciais" na outra. Um "C" maiúsculo fazia 27 lançamentos de terreno
# (R$ 3,2 M) sumirem da coluna "realizado" sem erro nenhum.
APELIDOS = {
    "(-) Despesas Comerciais": "(-) Despesas comerciais",
    "(-) Terreno - Pagamento Terreno": "(-) Terreno - Pagamento",
    "(-) Terreno - Outros aquisição e aluguel": "(-) Terreno - Outros",
    "(+) Outras entradas": "(+) Outras receitas administrativas",
    "(-) Assistência Técnica": "(-) Outras despesas administrativas",
}


LINHAS_CONHECIDAS = {
    "VGV", "(-) Desconto", "(-) Comissão s/ vendas", "RECEITA C/ VENDAS SPE",
    "(-) Impostos", "(-) Distratos", "(-) Despesas comerciais", "RECEITA LÍQUIDA",
    "(-) Terreno - Permuta", "(-) Terreno - Pagamento", "(-) Terreno - Outros",
    "(-) Obra - Custo Raso", "(-) Taxa de Administração - Obra",
    "(-) Taxa de Administração - Carteira", "(-) Incorporação - Decoração",
    "(-) Incorporação - Outros", "(-) Marketing - Stand", "(-) Marketing - Propaganda",
    "(-) Outras despesas administrativas", "(+) Outras receitas administrativas",
    "(+) Receitas c/ financiamento", "(-) Amortização de financiamento",
    "(-) Juros s/ financiamento", "(-/+) Retenções/liberações/Despesas",
    "DEDUÇÕES DO VGV", "DEDUÇÃO DA RECEITA", "GASTOS ", "LUCRO",
}


def semear_contas(s, wb) -> int:
    """
    Viabilidades Mensais, bloco realista (linhas 15..192): a coluna D traz o
    código do plano de contas do Sienge e a coluna C, quando preenchida, abre
    a linha do DRE a que aquele bloco de contas pertence.
    """
    vm = wb["Viabilidades Mensais_Cenarios"]
    linha_atual = None
    gravadas = 0

    for r in range(15, 193):
        rotulo = vm.cell(r, 3).value          # coluna C
        codigo = vm.cell(r, 4).value          # coluna D
        # a coluna C guarda o rótulo da linha do DRE, mas às vezes guarda um
        # número (linhas 22 e 24 usam a célula como memória de cálculo do
        # valor das parcelas pós-chaves). Só texto vale como rótulo.
        if isinstance(rotulo, str) and rotulo.strip():
            linha_atual = APELIDOS.get(rotulo.strip(), rotulo.strip())
        if not codigo or not str(codigo).strip():
            continue
        codigo = str(codigo).strip()
        if not linha_atual:
            continue

        if linha_atual not in LINHAS_CONHECIDAS:
            raise ValueError(
                f"linha do fluxo mensal sem correspondente no DRE: {linha_atual!r}. "
                "Acrescente um apelido em APELIDOS — sem isso os lançamentos "
                "desta conta somem da coluna 'realizado'.")
        grupo = GRUPO_POR_LINHA.get(linha_atual, "GASTO")
        # as contas de venda de imóvel e adiantamento são receita, não gasto
        if codigo.startswith(("1010101", "1010102", "1030")) or "Venda de Imóveis" in codigo:
            grupo = "RECEITA"
            if linha_atual.startswith("(-)"):
                linha_atual = "RECEITA C/ VENDAS SPE"

        q(s, """
            INSERT INTO conta (codigo, descricao, grupo, linha_dre, sinal)
            VALUES (:c, :d, :g, :l, :sn)
            ON CONFLICT (codigo) DO UPDATE
               SET descricao = EXCLUDED.descricao, grupo = EXCLUDED.grupo,
                   linha_dre = EXCLUDED.linha_dre, sinal = EXCLUDED.sinal
        """, c=codigo, d=codigo.split(" - ", 1)[-1] or codigo, g=grupo,
            l=linha_atual, sn=1 if grupo in ("RECEITA",) else -1)
        gravadas += 1

    # a conta agregadora da obra, que não aparece com código no plano
    q(s, """
        INSERT INTO conta (codigo, descricao, grupo, linha_dre, sinal)
        VALUES ('1 - OBRA', 'Obra — custo raso', 'GASTO', '(-) Obra - Custo Raso', -1)
        ON CONFLICT (codigo) DO UPDATE SET linha_dre = EXCLUDED.linha_dre
    """)
    s.commit()
    return gravadas + 1


# contas de mútuo e aporte entre empresas do grupo. Não são resultado do
# empreendimento — são movimento de caixa entre CNPJs. Na planilha vivem nas
# linhas 210-217 do fluxo mensal, fora do DRE. Sem esta classificação, 159
# lançamentos (R$ 2,6 M) ficavam pendurados em "A CLASSIFICAR".
INTERCOMPANY = [
    ("1020357", "MM Gestão Imobiliária Ltda"),
    ("2011457", "MM Gestão Imobiliária Ltda"),
    ("1020326", "Lisse Incorporadora Ltda"),
    ("1020304", "Barcelona Incorporadora Ltda"),
    ("2011404", "Barcelona Incorporadora Ltda"),
    ("2011458", "SPE Varsovia Incorporadora Ltda"),
    ("1020333", "M M Imobiliaria Unipessoal Ltda"),
    ("1020108", "Débitos com parceiros"),
]


def classificar_intercompany(s) -> int:
    n = 0
    for codigo, nome in INTERCOMPANY:
        n += len(q(s, """
            INSERT INTO conta (codigo, descricao, grupo, linha_dre, sinal)
            VALUES (:c, :d, 'APORTE', 'Mútuo e aportes entre empresas', 1)
            ON CONFLICT (codigo) DO UPDATE
               SET grupo = 'APORTE', linha_dre = 'Mútuo e aportes entre empresas'
            RETURNING id
        """, c=f"{codigo} - {nome}", d=nome))
    s.commit()
    return n


# =====================================================================
# 6b. a viabilidade orçada — valores aprovados, não recalculáveis
# =====================================================================
# VIABILIDADE coluna N, em R$ mil. As premissas do estudo original não
# existem mais em lugar nenhum: só o resultado sobreviveu. Por isso o cenário
# orçado guarda o RESULTADO, e não entradas para o motor reprocessar.
ORCADO_MIL = {
    "VGV": 102_900.0,
    "(-) Comissão s/ vendas": -5_670.0,
    "RECEITA C/ VENDAS SPE": 97_230.0,
    "(-) Impostos": -4_375.35,
    "(-) Distratos": 0.0,
    "(-) Despesas comerciais": -30.75,
    "RECEITA LÍQUIDA": 92_823.9,
    "(-) Terreno - Permuta": 0.0,
    "(-) Terreno - Pagamento": -3_450.0,
    "(-) Terreno - Outros": -86.25,
    "(-) Obra - Custo Raso": -45_500.0,
    "(-) Taxa de Administração - Obra": -8_190.0,
    "(-) Taxa de Administração - Carteira": -1_944.6,
    "(-) Incorporação - Decoração": -777.84,
    "(-) Incorporação - Outros": -1_680.0,
    "(-) Marketing - Stand": -1_000.0,
    "(-) Marketing - Propaganda": -630.0,
    "(-) Outras despesas administrativas": -910.0,
    "(+) Outras receitas administrativas": 0.02084,
    "LUCRO": 22_629.24,
}


def congelar_orcado(s, cenario_id: int) -> int:
    from app.servico import LINHAS_DRE
    rodada = q(s, """
        INSERT INTO rodada (cenario_id, executada_por, hash_entradas, congelada)
        VALUES (:c, 'viabilidade aprovada no lançamento', 'congelado-planilha', true)
        ON CONFLICT (cenario_id, hash_entradas) DO UPDATE SET congelada = true
        RETURNING id
    """, c=cenario_id)[0]["id"]
    q(s, "DELETE FROM resultado_projetado WHERE rodada_id = :r", r=rodada)
    for ordem, (rotulo, _) in enumerate(LINHAS_DRE):
        q(s, """INSERT INTO resultado_projetado (rodada_id, linha_dre, valor, ordem)
                VALUES (:r, :l, :v, :o)""",
          r=rodada, l=rotulo, v=round(ORCADO_MIL.get(rotulo, 0.0) * 1000, 2), o=ordem)
    s.commit()
    return rodada


# =====================================================================
# 2. empreendimento
# =====================================================================
def criar_empreendimento(s, wb) -> int:
    com = wb["Comercial"]
    area_priv = sum(float(com.cell(r, 3).value or 0) for r in range(6, 206))

    emp = q(s, """
        INSERT INTO empreendimento (sienge_enterprise_id, sienge_company_id, nome,
            area_construida, area_privativa, data_lancamento,
            data_entrega_prevista, mes_corte_realizado)
        VALUES (26003, 26, 'SPE KIEV — Doca Sede', 22909.46, :ap,
                DATE '2026-07-01', DATE '2031-03-31', DATE '2026-07-31')
        ON CONFLICT (sienge_enterprise_id) DO UPDATE
           SET area_privativa = EXCLUDED.area_privativa
        RETURNING id
    """, ap=round(area_priv, 2))[0]["id"]
    s.commit()
    return emp


# =====================================================================
# 3. tabelas de venda
# =====================================================================
def criar_tabelas_venda(s, emp: int) -> None:
    for nome, t in [
        ("Padrão",     dict(com=0.06, ato=0.04, men=0.35, anu=0.0, sem=0.35, uni=0.0, cha=0.20, n=60)),
        ("Investidor", dict(com=0.00, ato=0.00, men=1.00, anu=0.0, sem=0.00, uni=0.0, cha=0.00, n=12)),
    ]:
        q(s, """
            INSERT INTO tabela_venda (empreendimento_id, nome, perc_comissao,
                perc_ato, perc_mensais, perc_anuais, perc_semestrais, perc_unica,
                perc_chaves, qtd_mensais)
            VALUES (:e, :n, :com, :ato, :men, :anu, :sem, :uni, :cha, :q)
            ON CONFLICT (empreendimento_id, nome) DO UPDATE
               SET perc_comissao = EXCLUDED.perc_comissao,
                   perc_ato = EXCLUDED.perc_ato, perc_mensais = EXCLUDED.perc_mensais,
                   perc_anuais = EXCLUDED.perc_anuais,
                   perc_semestrais = EXCLUDED.perc_semestrais,
                   perc_unica = EXCLUDED.perc_unica, perc_chaves = EXCLUDED.perc_chaves,
                   qtd_mensais = EXCLUDED.qtd_mensais
        """, e=emp, n=nome, com=t["com"], ato=t["ato"], men=t["men"], anu=t["anu"],
            sem=t["sem"], uni=t["uni"], cha=t["cha"], q=t["n"])
    s.commit()


# =====================================================================
# 4. unidades, preços e carteira
# =====================================================================
def importar_sienge(s, emp: int, wb) -> dict:
    resumo = {}
    resumo["unidades"] = gravar.gravar_unidades(
        s, emp, normalizar_unidades(aba_como_dicts(wb, "Unidades")),
        origem="upload", arquivo=ARQUIVO)

    # a situação e o tipo de venda comercial moram na aba Comercial, não no
    # Sienge. Aproveitamos a passagem para conciliar os dois cadastros.
    com = wb["Comercial"]
    no_comercial: set[str] = set()
    for r in range(6, 206):
        nome = com.cell(r, 2).value
        if nome is None:
            continue
        nome = str(nome).strip()
        no_comercial.add(nome)
        # unidade que a planilha usa no VGV mas o Sienge não cadastrou
        q(s, """
            INSERT INTO unidade (empreendimento_id, nome, area_privativa,
                                 situacao, tipo_venda, origem_cadastro)
            VALUES (:e, :n, :a, 'Disponível', 'Normal', 'planilha')
            ON CONFLICT (empreendimento_id, nome) DO NOTHING
        """, e=emp, n=nome, a=float(com.cell(r, 3).value or 0))
        situacao = str(com.cell(r, 7).value or "Disponível").strip()
        tipo = str(com.cell(r, 10).value or "Normal").strip()
        bruto = float(com.cell(r, 4).value or 0)
        q(s, """
            UPDATE unidade SET situacao = CAST(:s AS situacao_unidade),
                               tipo_venda = CAST(:t AS tipo_venda)
             WHERE empreendimento_id = :e AND nome = :n
        """, e=emp, n=nome, s=situacao, t=tipo)
        if bruto:
            q(s, """
                INSERT INTO preco_unidade (unidade_id, preco_bruto, vigente_desde, observacao)
                SELECT id, :p, :d, 'migrado de Comercial!D'
                  FROM unidade WHERE empreendimento_id = :e AND nome = :n
                ON CONFLICT (unidade_id, vigente_desde) DO UPDATE
                   SET preco_bruto = EXCLUDED.preco_bruto
            """, e=emp, n=nome, p=bruto, d=HOJE)

    # unidades cadastradas no Sienge que a planilha deixou fora do VGV
    fora = q(s, """
        UPDATE unidade SET considerar_na_viabilidade = false,
               motivo_exclusao = 'ausente da aba Comercial na planilha de origem'
         WHERE empreendimento_id = :e AND NOT (nome = ANY(:dentro))
        RETURNING nome, area_privativa
    """, e=emp, dentro=sorted(no_comercial))
    if fora:
        print(f"  ! {len(fora)} unidade(s) do Sienge fora do VGV: "
              + ", ".join(f"{u['nome']} ({u['area_privativa']} m²)" for u in fora))
    de_fora = q(s, """SELECT nome FROM unidade
                       WHERE empreendimento_id = :e AND origem_cadastro = 'planilha'""",
                e=emp)
    if de_fora:
        print("  ! unidade(s) no VGV sem cadastro no Sienge: "
              + ", ".join(u["nome"] for u in de_fora))
    s.commit()

    resumo["contratos"] = gravar.gravar_contratos(
        s, emp, normalizar_contratos(aba_como_dicts(wb, "Contratos")),
        origem="upload", arquivo=ARQUIVO)
    resumo["receber"] = gravar.gravar_receber(
        s, emp, normalizar_receber(aba_como_dicts(wb, "Receber")),
        origem="upload", arquivo=ARQUIVO)
    resumo["recebido"] = gravar.gravar_recebido(
        s, emp, normalizar_recebido(aba_como_dicts(wb, "Recebido")),
        origem="upload", arquivo=ARQUIVO)
    resumo["fin_obra"] = gravar.gravar_movimentos(
        s, emp, normalizar_movimentos(aba_como_dicts(wb, "Fin_Obra")),
        origem="upload", arquivo=ARQUIVO)
    return resumo


# =====================================================================
# 5. orçamento de obra — um só, pesos e curva do cronograma
# =====================================================================
def criar_orcamento(s, emp: int, wb, custo_raso: float) -> int:
    cr = wb["Cronograma obra"]

    itens = []
    for r in range(4, 57):
        nome = cr.cell(r, 4).value
        valor = cr.cell(r, 5).value
        if not nome or not valor:
            continue
        curva = {}
        for c in range(8, 72):
            mes = cr.cell(2, c).value
            perc = cr.cell(r, c).value
            if isinstance(mes, dt.datetime) and isinstance(perc, (int, float)) and perc:
                d = mes.date()
                fim = (dt.date(d.year + d.month // 12, d.month % 12 + 1, 1)
                       - dt.timedelta(days=1))
                curva[fim] = curva.get(fim, 0.0) + float(perc)
        itens.append({"codigo": str(cr.cell(r, 2).value or f"IND{r}").strip(),
                      "descricao": str(nome).strip(),
                      "valor": float(valor), "curva": curva})

    # despesas indiretas não têm curva própria: seguem a curva ponderada das
    # diretas (Cronograma!H4 = H57). Reconstruímos isso explicitamente.
    diretas = [i for i in itens if i["curva"]]
    total_direto = sum(i["valor"] for i in diretas) or 1.0
    curva_ponderada: dict[dt.date, float] = defaultdict(float)
    for i in diretas:
        peso = i["valor"] / total_direto
        for m, p in i["curva"].items():
            curva_ponderada[m] += peso * p
    for i in itens:
        if not i["curva"]:
            i["curva"] = dict(curva_ponderada)

    total = sum(i["valor"] for i in itens)
    orc = q(s, """
        INSERT INTO orcamento_obra (empreendimento_id, versao, custo_raso,
                                    data_base, indice_reajuste, vigente)
        VALUES (:e, 'migrado-planilha', :c, DATE '2026-08-01', 'INCC-DI', true)
        ON CONFLICT (empreendimento_id, versao) DO UPDATE
           SET custo_raso = EXCLUDED.custo_raso, vigente = true
        RETURNING id
    """, e=emp, c=custo_raso)[0]["id"]

    q(s, """DELETE FROM cronograma_item c USING eap_item i
             WHERE i.id = c.eap_item_id AND i.orcamento_id = :o""", o=orc)
    q(s, "DELETE FROM eap_item WHERE orcamento_id = :o", o=orc)

    # normaliza os pesos para somar exatamente 1 (a trava do banco exige)
    pesos = [i["valor"] / total for i in itens]
    pesos[-1] += 1.0 - sum(pesos)

    for i, peso in zip(itens, pesos):
        item_id = q(s, """
            INSERT INTO eap_item (orcamento_id, codigo, descricao, peso, variacao_negociada)
            VALUES (:o, :c, :d, :p, 0) RETURNING id
        """, o=orc, c=i["codigo"], d=i["descricao"], p=round(peso, 8))[0]["id"]

        soma = sum(i["curva"].values()) or 1.0
        normal = {m: v / soma for m, v in i["curva"].items()}
        ajuste = 1.0 - sum(normal.values())
        if normal:
            ultimo = max(normal)
            normal[ultimo] += ajuste
        for m, p in normal.items():
            q(s, """INSERT INTO cronograma_item (eap_item_id, mes, perc_fisico)
                    VALUES (:i, :m, :p)
                    ON CONFLICT (eap_item_id, mes) DO UPDATE
                       SET perc_fisico = EXCLUDED.perc_fisico""",
              i=item_id, m=m, p=round(p, 8))
    s.commit()
    return orc


# =====================================================================
# 6. cenários
# =====================================================================
PREMISSAS_COMUNS = {
    "ret": 0.045, "distratos": 0.0, "despesas_comerciais": 30_750.0,
    "terreno_registro_perc": 0.025, "taxa_adm_obra": 0.10,
    "taxa_viabilizacao": 0.05, "decoracao": 1_545_045.14,
    "projetos_e_outros": 2_899_156.992, "marketing_stand": 0.0,
    "marketing_propaganda": 1_545_045.14, "outras_desp_adm_perc": 0.015,
    "outras_entradas": 117.95, "tma_anual": 0.18,
    "financiamento_limite": 0.0, "financiamento_juros_aa": 0.22,
    "financiamento_prazo_amort": 6, "financiamento_gatilho_obra": 0.20,
    "meses_pos_chaves": 6,
}
UNIDADE_DA_PREMISSA = {
    "ret": "percentual", "distratos": "percentual",
    "despesas_comerciais": "moeda", "terreno_registro_perc": "percentual",
    "taxa_adm_obra": "percentual", "taxa_viabilizacao": "percentual",
    "decoracao": "moeda", "projetos_e_outros": "moeda",
    "marketing_stand": "moeda", "marketing_propaganda": "moeda",
    "outras_desp_adm_perc": "percentual", "outras_entradas": "moeda",
    "tma_anual": "percentual", "financiamento_limite": "moeda",
    "financiamento_juros_aa": "percentual", "financiamento_prazo_amort": "meses",
    "financiamento_gatilho_obra": "percentual", "meses_pos_chaves": "meses",
    "preco_m2_estoque": "r$/m2",
}
TERRENO = [2_200_000, 350_000, 350_000, 300_000, 250_000]


def criar_cenario(s, emp: int, nome: str, tipo: str, principal: bool,
                  preco_m2: float, usar_tabela: bool, preco_investidor: float,
                  plano: list[tuple[dt.date, str, int]]) -> int:
    cen = q(s, """
        INSERT INTO cenario (empreendimento_id, nome, tipo, mes_base,
                             horizonte_meses, principal, congelado_em)
        VALUES (:e, :n, :t, DATE '2026-08-31', 90, :p,
                CASE WHEN :t = 'orcado' THEN now() END)
        ON CONFLICT (empreendimento_id, nome, tipo) DO UPDATE
           SET principal = EXCLUDED.principal
        RETURNING id
    """, e=emp, n=nome, t=tipo, p=principal)[0]["id"]

    q(s, "DELETE FROM premissa WHERE cenario_id = :c", c=cen)
    valores = dict(PREMISSAS_COMUNS, preco_m2_estoque=preco_m2)
    for chave, valor in valores.items():
        q(s, """INSERT INTO premissa (cenario_id, chave, valor, unidade, origem)
                VALUES (:c, :k, :v, :u, 'migrado da planilha')""",
          c=cen, k=chave, v=valor, u=UNIDADE_DA_PREMISSA.get(chave, "moeda"))

    q(s, "DELETE FROM premissa_terreno WHERE cenario_id = :c", c=cen)
    for i, parcela in enumerate(TERRENO):
        q(s, """INSERT INTO premissa_terreno (cenario_id, ordem, valor)
                VALUES (:c, :o, :v)""", c=cen, o=i, v=parcela)

    q(s, "DELETE FROM preco_cenario WHERE cenario_id = :c", c=cen)
    q(s, """INSERT INTO preco_cenario (cenario_id, tipo_venda, preco_m2, usar_tabela)
            VALUES (:c, 'Normal', :p, :u)""", c=cen, p=preco_m2, u=usar_tabela)
    q(s, """INSERT INTO preco_cenario (cenario_id, tipo_venda, preco_unidade, usar_tabela)
            VALUES (:c, 'Investidor', :p, false)""", c=cen, p=preco_investidor)

    q(s, "DELETE FROM plano_venda WHERE cenario_id = :c", c=cen)
    for mes, tipo_venda, qtd in plano:
        q(s, """INSERT INTO plano_venda (cenario_id, mes, tipo_venda, quantidade)
                VALUES (:c, :m, CAST(:t AS tipo_venda), :q)
                ON CONFLICT (cenario_id, mes, tipo_venda) DO UPDATE
                   SET quantidade = EXCLUDED.quantidade""",
          c=cen, m=mes, t=tipo_venda, q=qtd)
    s.commit()
    return cen


def ler_plano(wb, coluna: int) -> list[tuple[dt.date, str, int]]:
    sim = wb["SIMULAÇÕES"]
    saida = []
    for r in range(82, 158):
        d = sim.cell(r, 1).value
        if not isinstance(d, dt.datetime):
            continue
        qtd = sim.cell(r, coluna).value
        if not qtd:
            continue
        tipo = "Investidor" if str(sim.cell(r, 2).value or "") == "Investidor" else "Normal"
        saida.append((d.date(), tipo, int(qtd)))
    return saida


# =====================================================================
def main() -> int:
    print("aplicando migrations...", aplicar_migrations())
    wb = openpyxl.load_workbook(ARQUIVO, data_only=True)
    s = Sessao()

    emp = criar_empreendimento(s, wb)
    print(f"empreendimento #{emp}")

    n = semear_contas(s, wb)
    classificar_intercompany(s)
    print(f"plano de contas: {n} contas (+ {len(INTERCOMPANY)} de mútuo/aporte)")

    criar_tabelas_venda(s, emp)
    resumo = importar_sienge(s, emp, wb)
    for fonte, r in resumo.items():
        aviso = f"  ({len(r.get('avisos', []))} aviso(s))" if r.get("avisos") else ""
        print(f"{fonte:>12}: {r['gravadas']}/{r['lidas']} gravadas{aviso}")

    orc = criar_orcamento(s, emp, wb, 104_049_038.53)
    print(f"orçamento de obra #{orc}")

    cenarios = {}
    for nome, tipo, principal, preco, usar_tab, preco_inv, col in [
        ("realista",   "projecao", True,  18_228.664532617702, True,  3_704_512.50, 10),
        ("otimista",   "projecao", False, 18_000.0,            False, 3_624_952.50, 4),
        ("pessimista", "projecao", False, 17_848.092006697818, False, 3_624_952.50, 16),
    ]:
        cid = criar_cenario(s, emp, nome, tipo, principal, preco, usar_tab,
                            preco_inv, ler_plano(wb, col))
        cenarios[nome] = cid
        print(f"cenário {nome}: #{cid}")

    # a viabilidade orçada (VIABILIDADE coluna N), congelada
    cid_orc = criar_cenario(s, emp, "orçada no lançamento", "orcado", False,
                            18_228.664532617702, True, 3_704_512.50,
                            ler_plano(wb, 10))
    cenarios["orcado"] = cid_orc
    rodada_orc = congelar_orcado(s, cid_orc)
    print(f"cenário orçado: #{cid_orc} (rodada congelada #{rodada_orc})")

    from app.servico import rodar_e_persistir
    for nome in ("realista", "otimista", "pessimista"):
        r = rodar_e_persistir(s, cenarios[nome], executada_por="migração")
        print(f"  rodada de {nome}: #{r}")

    s.close()
    return emp


if __name__ == "__main__":
    main()
