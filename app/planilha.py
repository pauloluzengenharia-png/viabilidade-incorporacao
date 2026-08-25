"""
A memória de cálculo de uma linha, em planilha.

Existe para o número sair do sistema com a conta junto: anexar numa ata, mandar
para o contábil, conferir fora. Duas regras que a planilha respeita e a antiga
não respeitava:

**Os resultados são fórmulas, não valores colados.** A aba `Fórmula` calcula o
resultado a partir dos operandos com uma fórmula de verdade; a de lançamentos
soma com `SUM`. Quem receber pode mexer num operando e ver o efeito — que é o
que se espera de uma planilha, e exatamente o que a planilha de origem tinha
deixado de fazer em várias linhas.

**Cada premissa aparece com a sua origem.** Ao lado do valor fica de onde ele
veio e, quando alguém o editou, quem e quando.
"""
from __future__ import annotations

import datetime as dt
import io
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONTE = "Arial"
TINTA = "FF191C1A"
AREIA = "FFE0D0BF"
PRETO = "FF000000"
CINZA = "FF726C64"
AZUL_ENTRADA = "FF0000FF"          # convenção: azul = valor digitado

MOEDA = '#,##0.00;(#,##0.00);-'
PERCENT = '0.0000%'
DATA = 'DD/MM/YYYY'
MES = 'MM/YYYY'

_fina = Side(style="thin", color="FFDFDAD2")
BORDA = Border(bottom=_fina)


def _titulo(ws, linha: int, texto: str, largura: int = 6) -> int:
    c = ws.cell(row=linha, column=1, value=texto)
    c.font = Font(name=FONTE, size=12, bold=True, color=TINTA)
    return linha + 1


def _cabecalho(ws, linha: int, colunas: list[str]) -> int:
    for i, nome in enumerate(colunas, start=1):
        c = ws.cell(row=linha, column=i, value=nome)
        c.font = Font(name=FONTE, size=9, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=PRETO)
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[linha].height = 20
    return linha + 1


def _larguras(ws, larguras: list[int]) -> None:
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _texto(ws, linha: int, col: int, valor: Any, *, negrito=False, cor=TINTA,
           formato: Optional[str] = None, tamanho=10):
    c = ws.cell(row=linha, column=col, value=valor)
    c.font = Font(name=FONTE, size=tamanho, bold=negrito, color=cor)
    if formato:
        c.number_format = formato
    c.border = BORDA
    return c


def memoria_de_calculo(*, emp: dict, cenario: dict, rotulo: str,
                       corte: Optional[dt.date], formula, insumos: dict,
                       realizado: dict, projecao: list, linha_tabela: dict,
                       verbete) -> bytes:
    wb = Workbook()

    # ---------------------------------------------------------- resumo
    ws = wb.active
    ws.title = "Resumo"
    _larguras(ws, [34, 26, 26, 26, 26])

    l = _titulo(ws, 1, f"{rotulo} — memória de cálculo")
    _texto(ws, l, 1, emp["nome"], tamanho=11); l += 1
    _texto(ws, l, 1, f"Cenário {cenario['nome']} · realizado até "
                     f"{corte:%m/%Y}" if corte else f"Cenário {cenario['nome']}",
           cor=CINZA, tamanho=9)
    l += 1
    _texto(ws, l, 1, f"Gerado pelo sistema de viabilidade da MMI em "
                     f"{dt.datetime.now():%d/%m/%Y %H:%M}", cor=CINZA, tamanho=9)
    l += 2

    if verbete and verbete.o_que:
        c = _texto(ws, l, 1, verbete.o_que, cor=CINZA, tamanho=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=l, start_column=1, end_row=l, end_column=5)
        ws.row_dimensions[l].height = 30
        l += 2

    if linha_tabela:
        l = _cabecalho(ws, l, ["", "Orçado", "Atualizado", "Realizado", "A realizar"])
        _texto(ws, l, 1, "Valor (R$)", negrito=True)
        for col, chave in enumerate(("orcado", "atualizado", "realizado"), start=2):
            _texto(ws, l, col, linha_tabela[chave], formato=MOEDA)
        # a última coluna é conta, não número colado
        _texto(ws, l, 5, f"=C{l}-D{l}", formato=MOEDA)
        l += 1
        _texto(ws, l, 1, "De onde vem", cor=CINZA, tamanho=9)
        for col, txt in enumerate(("congelado no lançamento",
                                   f"cenário {cenario['nome']}, premissas de hoje",
                                   "lançamentos do Sienge",
                                   "atualizado − realizado"), start=2):
            _texto(ws, l, col, txt, cor=CINZA, tamanho=9)

    # ---------------------------------------------------------- fórmula
    if formula:
        ws = wb.create_sheet("Fórmula")
        _larguras(ws, [40, 22, 52, 30])
        l = _titulo(ws, 1, "Como o valor atualizado é calculado")
        _texto(ws, l, 1, formula.texto, negrito=True); l += 2

        l = _cabecalho(ws, l, ["Operando", "Valor", "De onde vem", "Última alteração"])
        primeira = l
        for termo in formula.termos:
            _texto(ws, l, 1, termo.rotulo)
            c = _texto(ws, l, 2, termo.valor,
                       formato=PERCENT if termo.formato == "percentual" else MOEDA,
                       cor=AZUL_ENTRADA)
            origem = termo.origem
            info = insumos.get(termo.chave) if termo.chave else None
            if info:
                origem = f"{origem} ({info['origem']})" if origem else info["origem"]
            _texto(ws, l, 3, origem, cor=CINZA, tamanho=9)
            if info and info.get("alteracao"):
                a = info["alteracao"]
                _texto(ws, l, 4,
                       f"{a['autor']} em {a['em']:%d/%m/%Y}: "
                       f"{a['valor_anterior']} → {a['valor_novo']}",
                       cor=CINZA, tamanho=9)
            l += 1
        ultima = l - 1

        # o resultado é fórmula de verdade: mexer num operando recalcula
        excel = formula.excel.replace("{todos}", f"B{primeira}:B{ultima}")
        for i in range(1, len(formula.termos) + 1):
            excel = excel.replace(f"{{{i}}}", f"B{primeira + i - 1}")
        _texto(ws, l, 1, "Resultado", negrito=True)
        c = _texto(ws, l, 2, excel, negrito=True, formato=MOEDA)
        _texto(ws, l, 3, "é o número da coluna Atualizado", cor=CINZA, tamanho=9)
        l += 2

        _texto(ws, l, 1, "Os valores em azul são os operandos. Mexer num deles "
                         "recalcula o resultado — a planilha não traz o número "
                         "colado.", cor=CINZA, tamanho=9)
        l += 1
        if formula.nota:
            c = _texto(ws, l, 1, f"Vale saber: {formula.nota}", cor=CINZA, tamanho=9)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=l, start_column=1, end_row=l, end_column=4)
            ws.row_dimensions[l].height = 42

    # ---------------------------------------------------- lançamentos
    if realizado.get("linhas"):
        ws = wb.create_sheet("Realizado")
        if realizado["tipo"] == "comissao":
            _larguras(ws, [22, 16, 20, 20, 20])
            l = _titulo(ws, 1, "Contratos que compõem a comissão realizada")
            if realizado.get("nota"):
                c = _texto(ws, l, 1, realizado["nota"], cor=CINZA, tamanho=9)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(start_row=l, start_column=1, end_row=l, end_column=5)
                ws.row_dimensions[l].height = 46
            l += 2
            l = _cabecalho(ws, l, ["Unidade", "Contrato", "Preço de tabela",
                                   "Líquido do contrato", "Comissão"])
            primeira = l
            for m in realizado["linhas"]:
                _texto(ws, l, 1, m["nome"])
                _texto(ws, l, 2, m["data_contrato"], formato=DATA)
                _texto(ws, l, 3, float(m["preco_bruto"]), formato=MOEDA)
                _texto(ws, l, 4, float(m["liquido"]), formato=MOEDA)
                _texto(ws, l, 5, f"=D{l}-C{l}", formato=MOEDA)
                l += 1
            _texto(ws, l, 1, f"{len(realizado['linhas'])} contrato(s)", negrito=True)
            _texto(ws, l, 5, f"=SUM(E{primeira}:E{l - 1})", negrito=True, formato=MOEDA)
        else:
            _larguras(ws, [13, 13, 18, 34, 34, 12, 20])
            l = _titulo(ws, 1, "Lançamentos do Sienge que somam o realizado")
            l += 1
            l = _cabecalho(ws, l, ["Data", "Competência", "Conta", "Descrição",
                                   "Fornecedor", "Rateio", "Valor"])
            primeira = l
            for m in realizado["linhas"]:
                _texto(ws, l, 1, m["data_movimento"], formato=DATA)
                _texto(ws, l, 2, m["mes_competencia"], formato=MES)
                _texto(ws, l, 3, m["conta_codigo"])
                _texto(ws, l, 4, m["conta"])
                _texto(ws, l, 5, m["fornecedor"] or "")
                r = m["rateio_categoria"]
                _texto(ws, l, 6, float(r) / 100 if r is not None else 1,
                       formato='0.00%')
                _texto(ws, l, 7, float(m["valor"]), formato=MOEDA)
                l += 1
            _texto(ws, l, 1, f"{len(realizado['linhas'])} lançamento(s)", negrito=True)
            _texto(ws, l, 7, f"=SUM(G{primeira}:G{l - 1})", negrito=True, formato=MOEDA)
            l += 2
            _texto(ws, l, 1, "Um pagamento rateado entre categorias aparece mais "
                             "de uma vez, cada linha com a sua fração — o rateio "
                             "está na coluna Rateio.", cor=CINZA, tamanho=9)

    # ------------------------------------------------------- projeção
    if projecao:
        ws = wb.create_sheet("Projeção")
        _larguras(ws, [16, 22, 22])
        l = _titulo(ws, 1, f"Distribuição no tempo — cenário {cenario['nome']}")
        l += 1
        l = _cabecalho(ws, l, ["Mês", "Valor projetado", "Acumulado"])
        primeira = l
        for item in projecao:
            _texto(ws, l, 1, item["mes"], formato=MES)
            _texto(ws, l, 2, float(item["valor"]), formato=MOEDA)
            _texto(ws, l, 3, f"=SUM($B${primeira}:B{l})", formato=MOEDA)
            l += 1
        _texto(ws, l, 1, "Total", negrito=True)
        _texto(ws, l, 2, f"=SUM(B{primeira}:B{l - 1})", negrito=True, formato=MOEDA)
        l += 2
        c = _texto(ws, l, 1,
                   "O total daqui pode não bater com a linha do resultado por "
                   "três motivos, todos esperados: parte do valor cai fora do "
                   "horizonte projetado; a linha inclui permuta, que é custo e "
                   "não é caixa; ou cada mês foi arredondado a centavos na "
                   "gravação, o que numa obra de anos dá alguns reais no total.",
                   cor=CINZA, tamanho=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=l, start_column=1, end_row=l, end_column=3)
        ws.row_dimensions[l].height = 42

    for aba in wb.worksheets:
        aba.sheet_view.showGridLines = False
        aba.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
