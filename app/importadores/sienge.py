"""
Importador dos dados do Sienge.

Camada adaptadora: hoje a fonte é um arquivo exportado; amanhã é a API REST.
As funções `normalizar_*` recebem linhas cruas (lista de dicts, do jeito que
saem tanto do export quanto do JSON da API) e devolvem registros prontos para
o banco. Trocar upload por API não encosta no motor nem no serviço.

As três regras não óbvias que vieram da planilha e moram aqui:

  1. `item de orçamento` — Fin_Obra!BB. Casa o lançamento bancário com a linha
     do DRE de viabilidade. Sem isso o realizado não tem onde cair.
  2. `valor rateado` — Fin_Obra!BC. O valor já entra rateado por categoria
     financeira e por departamento. Guardar o bruto é o caminho mais curto
     para contar duas vezes.
  3. `condição` da parcela — Receber!BG. 1 = poupança (mensais até as chaves),
     2 e 3 = pós-chaves (repasse/financiamento). É o que separa os dois
     regimes de recebimento no fluxo de caixa.
"""
from __future__ import annotations

import datetime as dt
import math
from typing import Any, Iterable, Optional

Linha = dict[str, Any]


# ---------------------------------------------------------------- utilidades
def _txt(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v: Any, padrao: float = 0.0) -> float:
    if v is None or v == "":
        return padrao
    if isinstance(v, (int, float)):
        return padrao if isinstance(v, float) and math.isnan(v) else float(v)
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return padrao


def _data(v: Any) -> Optional[dt.date]:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()[:10]
    for f in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(s, f).date()
        except ValueError:
            continue
    return None


def fim_do_mes(d: dt.date) -> dt.date:
    proximo = dt.date(d.year + d.month // 12, d.month % 12 + 1, 1)
    return proximo - dt.timedelta(days=1)


def _campo(l: Linha, *nomes: str) -> Any:
    """Aceita tanto o nome do export quanto o da API (que muda de prefixo)."""
    for n in nomes:
        if n in l and l[n] not in (None, ""):
            return l[n]
    # tolera o prefixo 'Value.' / 'Column1.' que o Power Query cria
    for chave, valor in l.items():
        curto = str(chave).split(".")[-1]
        if curto in nomes and valor not in (None, ""):
            return valor
    return None


# ---------------------------------------------------------------- unidades
SITUACAO_SIENGE = {
    "V": "Vendida", "D": "Disponível", "R": "Reservada", "P": "Permuta",
    "Vendida": "Vendida", "Disponível": "Disponível", "Disponivel": "Disponível",
    "Reservada": "Reservada", "Permuta": "Permuta", "Distratada": "Distratada",
}


def normalizar_unidades(linhas: Iterable[Linha]) -> list[Linha]:
    saida = []
    for l in linhas:
        nome = _txt(_campo(l, "name", "Column1.units.name"))
        if not nome or nome.lower() in {"name", "nome", "unidade"}:
            continue  # cabeçalho repetido de uma 2ª tabela na mesma aba
        area = _campo(l, "privateArea")
        if area is not None and not isinstance(area, (int, float)):
            continue  # linha de cabeçalho disfarçada
        bruta = _campo(l, "Situação", "commercialStock", "situacao")
        saida.append({
            "sienge_unit_id": int(_num(_campo(l, "id"), 0)) or None,
            "nome": nome,
            "tipo_imovel": _txt(_campo(l, "propertyType")),
            "area_privativa": _num(_campo(l, "privateArea")),
            "fracao_ideal": _num(_campo(l, "idealFraction")) or None,
            "situacao": SITUACAO_SIENGE.get(_txt(bruta) or "", "Disponível"),
        })
    return saida


# ---------------------------------------------------------------- contratos
def normalizar_contratos(linhas: Iterable[Linha]) -> list[Linha]:
    saida = []
    for l in linhas:
        unidade = _txt(_campo(l, "Column1.units.name", "name"))
        if not unidade:
            continue
        situacao = _txt(_campo(l, "Column1.situation", "situation")) or ""
        if situacao.upper().startswith("CANCEL"):
            continue
        saida.append({
            "unidade": unidade,
            "sienge_contract_id": int(_num(_campo(l, "Column1.id"), 0)) or None,
            "sienge_bill_id": int(_num(_campo(l, "Column1.receivableBillId",
                                              "receivableBillId"), 0)) or None,
            "cliente_nome": (_txt(_campo(l, "Column1.customers.name",
                                         "customers.name")) or "").upper() or None,
            "data_contrato": _data(_campo(l, "Column1.contractDate", "contractDate")),
            "valor_total": _num(_campo(l, "Column1.value", "value")),
            "valor_venda": _num(_campo(l, "Column1.totalSellingValue", "totalSellingValue")),
            "indexador": _txt(_campo(l, "Column1.correctionType", "correctionType")),
            "situacao": situacao or None,
        })
    return saida


# ---------------------------------------------------------------- carteira
def normalizar_receber(linhas: Iterable[Linha]) -> list[Linha]:
    """Receber → parcela_receber. `Condição` separa poupança de pós-chaves."""
    saida = []
    for l in linhas:
        unidade = _txt(_campo(l, "Unidade", "mainUnit"))
        valor = _num(_campo(l, "Valor", "balanceAmount", "correctedBalanceAmount"))
        venc = _data(_campo(l, "Data", "dueDate"))
        if not unidade or not venc or not valor:
            continue
        saida.append({
            "unidade": unidade,
            "sienge_bill_id": int(_num(_campo(l, "billId"), 0)) or None,
            "sienge_installment_id": int(_num(_campo(l, "installmentId"), 0)) or None,
            "numero_parcela": _txt(_campo(l, "installmentNumber")),
            "condicao": int(_num(_campo(l, "Condição", "Condicao"), 1)) or 1,
            "vencimento": venc,
            "valor": valor,
            "indexador": _txt(_campo(l, "indexerName")),
        })
    return saida


def normalizar_recebido(linhas: Iterable[Linha]) -> list[Linha]:
    saida = []
    for l in linhas:
        unidade = _txt(_campo(l, "Unidade", "mainUnit"))
        valor = _num(_campo(l, "Valor", "receipts.netAmount"))
        quando = _data(_campo(l, "Recebimento", "receipts.paymentDate", "dueDate"))
        if not unidade or not quando or not valor:
            continue
        saida.append({
            "unidade": unidade,
            "sienge_bill_id": int(_num(_campo(l, "billId"), 0)) or None,
            "sienge_installment_id": int(_num(_campo(l, "installmentId"), 0)) or None,
            "data_recebimento": quando,
            "valor_liquido": valor,
            "empresa_id": int(_num(_campo(l, "companyId"), 0)) or None,
        })
    return saida


# ---------------------------------------------------------------- incorrido
UNIDADES_DE_OBRA = {"OBRA", "CUSTOS INDIRETOS - ORÇAMENTO OBRA"}
CODIGO_OBRA = "1 - OBRA"


def item_de_orcamento(l: Linha) -> str:
    """
    Fin_Obra!BB, traduzido:

        se a unidade construtiva é OBRA (ou custos indiretos da obra)
            → tudo cai na conta única "1 - OBRA"
        senão, se o lançamento não tem planilha de custo
            → usa a categoria financeira
        senão
            → usa o item da planilha de custo
    """
    unidade_constr = (_txt(_campo(l, "buildingUnitName")) or "").upper()
    if unidade_constr in UNIDADES_DE_OBRA:
        return CODIGO_OBRA

    planilha_id = _txt(_campo(l, "costEstimationSheetId"))
    planilha_nome = _txt(_campo(l, "costEstimationSheetName"))
    if planilha_id or planilha_nome:
        return f"{planilha_id or ''} - {planilha_nome or ''}"

    cat_id = _txt(_campo(l, "financialCategoryId"))
    cat_nome = _txt(_campo(l, "financialCategoryName"))
    return f"{cat_id or ''} - {cat_nome or ''}"


def valor_rateado(l: Linha) -> float:
    """
    Fin_Obra!BC: valor × rateio da categoria × rateio do departamento.

    Cuidado com as duas escalas, que são diferentes e não estão documentadas
    em lugar nenhum da planilha:

        financialCategoryRate  vem em PONTOS PERCENTUAIS  (100 = integral, 50 = metade)
        %  (departamentCosts)  vem em FRAÇÃO             (1   = integral, 0.5 = metade)

    Na base da Kiev: 764 lançamentos com categoria 100 e 4 com 50; 754 com
    departamento 1 e 14 com 0,5. Tratar os dois na mesma escala erra 18 linhas.

    O sinal já vem embutido em `Personalizar` (negativo = saída de caixa),
    e não em `bankMovementAmount`, que é sempre positivo.
    """
    bruto = _num(_campo(l, "Personalizar"))
    if not bruto:
        # sem a coluna personalizada, cai no valor bruto — sem sinal confiável
        bruto = _num(_campo(l, "bankMovementAmount", "amount"))
    rateio_cat = _num(_campo(l, "financialCategoryRate"), 100.0)      # 0..100
    rateio_dep = _num(_campo(l, "%", "departamentCosts"), 1.0)        # 0..1
    return bruto * (rateio_cat / 100.0) * rateio_dep


def normalizar_movimentos(linhas: Iterable[Linha]) -> list[Linha]:
    vistos: dict[tuple, int] = {}
    saida = []
    for l in linhas:
        quando = _data(_campo(l, "bankMovementDate", "Fim Mês"))
        if not quando:
            continue
        valor = _num(_campo(l, "Valor")) or valor_rateado(l)
        if not valor:
            continue
        bank_id = int(_num(_campo(l, "bankMovementId"), 0)) or None
        conta = _txt(_campo(l, "Item de Orçamento")) or item_de_orcamento(l)
        chave = (bank_id, conta)
        vistos[chave] = vistos.get(chave, 0) + 1
        saida.append({
            "sienge_bank_movement_id": bank_id,
            "sequencia": vistos[chave],
            "conta_codigo": conta,
            "data_movimento": quando,
            "valor": valor,
            "rateio_categoria": _num(_campo(l, "financialCategoryRate"), 100.0),
            "rateio_departamento": _num(_campo(l, "%", "departamentCosts"), 100.0),
            "unidade": _txt(_campo(l, "buildingUnitName")),
            "fornecedor": _txt(_campo(l, "creditorName")),
            "centro_custo": _txt(_campo(l, "costCenterName")),
            "conciliado": str(_campo(l, "bankMovementReconcile") or "").upper() == "S",
        })
    return saida


# ---------------------------------------------------------------- leitura xlsx
def ler_planilha(caminho_ou_buffer, aba: Optional[str] = None,
                 linha_cabecalho: int = 1) -> list[Linha]:
    """Lê uma aba de xlsx como lista de dicts, usando a 1ª linha como cabeçalho."""
    import openpyxl
    wb = openpyxl.load_workbook(caminho_ou_buffer, data_only=True, read_only=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]

    linhas = ws.iter_rows(values_only=True)
    for _ in range(linha_cabecalho - 1):
        next(linhas, None)
    cabecalho = next(linhas, None)
    if not cabecalho:
        return []
    cols = [str(c).strip() if c is not None else f"col{i}"
            for i, c in enumerate(cabecalho)]

    saida = []
    for valores in linhas:
        if all(v is None for v in valores):
            continue
        saida.append({cols[i]: v for i, v in enumerate(valores) if i < len(cols)})
    wb.close()
    return saida
